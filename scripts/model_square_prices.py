#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""导出模型广场中全部模型的四类 Token 单价。

价格计算与模型广场保持相同口径：默认采用当前用户可见分组中的最低倍率，
并使用 /api/status 返回的美元兑人民币汇率换算为元/百万 Token。

用法：
    python scripts/model_square_prices.py --token <控制台访问令牌>
    python scripts/model_square_prices.py --token <令牌> --multiplier 0.8
    python scripts/model_square_prices.py --base-url https://example.com --output prices.xlsx

官方价格在同目录 official_model_prices.json 中维护，单位同样为元/百万 Token。
倍率只应用于该文件中已配置官方价格的模型；未配置的模型保持模型广场原价。
相对于官网的折扣仅按双方的非缓存输入单价计算。
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Any
from urllib.error import HTTPError, URLError
from urllib.request import Request, urlopen


SCRIPT_DIR = Path(__file__).resolve().parent
DEFAULT_OFFICIAL_PRICES = SCRIPT_DIR / "official_model_prices.json"
DEFAULT_OUTPUT = SCRIPT_DIR / "model_square_prices.xlsx"
PRICE_FIELDS = ("input", "output", "cache_read", "cache_write")
TABLE_HEADERS = (
    "模型",
    "输入(元/百万 Token)",
    "输出(元/百万 Token)",
    "缓存读取(元/百万 Token)",
    "缓存写入(元/百万 Token)",
    "相对于官网折扣",
)
TABLE_NOTES = (
    "1. 按量收费，使用多少扣多少",
    "2. GPT 模型，北京专线优化，首字延迟低",
)
NUMBER_PATTERN = r"[+-]?(?:\d+(?:\.\d*)?|\.\d+)(?:[eE][+-]?\d+)?"
TIER_PATTERN = re.compile(r'tier\("[^"]*",\s*([^)]+)\)')
COEFFICIENT_PATTERN = re.compile(
    rf"\b(p|c|cr|cc)\s*\*\s*({NUMBER_PATTERN})"
)
DYNAMIC_FIELD_MAP = {
    "p": "input",
    "c": "output",
    "cr": "cache_read",
    "cc": "cache_write",
}


def positive_decimal(value: str) -> Decimal:
    """解析命令行中的正数。"""
    try:
        result = Decimal(value)
    except InvalidOperation as exc:
        raise argparse.ArgumentTypeError(f"不是有效数字：{value}") from exc
    if not result.is_finite() or result <= 0:
        raise argparse.ArgumentTypeError("必须是大于 0 的有限数字")
    return result


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="导出模型广场全部模型的四类人民币单价",
    )
    parser.add_argument(
        "--base-url",
        default=os.getenv("MODEL_SQUARE_BASE_URL", "http://localhost:3000"),
        help="站点地址，默认读取 MODEL_SQUARE_BASE_URL 或 http://localhost:3000",
    )
    parser.add_argument(
        "--token",
        default=os.getenv("MODEL_SQUARE_TOKEN", ""),
        help="控制台访问令牌；也可通过 MODEL_SQUARE_TOKEN 设置",
    )
    parser.add_argument(
        "--official-prices",
        type=Path,
        default=DEFAULT_OFFICIAL_PRICES,
        help=f"官方价格 JSON，默认 {DEFAULT_OFFICIAL_PRICES}",
    )
    parser.add_argument(
        "--multiplier",
        type=positive_decimal,
        default=Decimal("1"),
        help="已配置官方价格模型的单价倍率，默认 1",
    )
    parser.add_argument(
        "--timeout",
        type=positive_decimal,
        default=Decimal("15"),
        help="HTTP 请求超时秒数，默认 15",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=DEFAULT_OUTPUT,
        help=f"输出 Excel 文件，默认 {DEFAULT_OUTPUT}",
    )
    return parser.parse_args()


def request_json(base_url: str, path: str, token: str, timeout: float) -> dict[str, Any]:
    url = f"{base_url.rstrip('/')}{path}"
    headers = {"Accept": "application/json"}
    if token.strip():
        headers["Authorization"] = f"Bearer {token.strip()}"
    request = Request(url, headers=headers)
    try:
        with urlopen(request, timeout=timeout) as response:
            payload = json.load(response)
    except HTTPError as exc:
        detail = exc.read().decode("utf-8", errors="replace")
        raise RuntimeError(f"请求 {url} 失败：HTTP {exc.code}，{detail}") from exc
    except (URLError, TimeoutError) as exc:
        raise RuntimeError(f"请求 {url} 失败：{exc}") from exc
    except json.JSONDecodeError as exc:
        raise RuntimeError(f"{url} 未返回有效 JSON") from exc

    if not isinstance(payload, dict):
        raise RuntimeError(f"{url} 返回的数据不是 JSON 对象")
    if payload.get("success") is not True:
        raise RuntimeError(f"请求 {url} 失败：{payload.get('message') or '未知错误'}")
    return payload


def to_decimal(value: Any, field_name: str) -> Decimal:
    if isinstance(value, bool):
        raise ValueError(f"{field_name} 不能是布尔值")
    try:
        result = Decimal(str(value))
    except (InvalidOperation, ValueError) as exc:
        raise ValueError(f"{field_name} 不是有效数字：{value!r}") from exc
    if not result.is_finite():
        raise ValueError(f"{field_name} 必须是有限数字")
    return result


def load_official_prices(path: Path) -> dict[str, dict[str, Decimal | None]]:
    try:
        with path.open("r", encoding="utf-8") as file:
            raw = json.load(file)
    except FileNotFoundError as exc:
        raise RuntimeError(f"官方价格文件不存在：{path}") from exc
    except json.JSONDecodeError as exc:
        raise RuntimeError(
            f"官方价格文件 JSON 格式错误：{path}:{exc.lineno}:{exc.colno}"
        ) from exc

    if not isinstance(raw, dict):
        raise RuntimeError("官方价格 JSON 顶层必须是对象")

    prices: dict[str, dict[str, Decimal | None]] = {}
    for model_name, model_prices in raw.items():
        if model_name.startswith("_"):
            continue
        if not isinstance(model_prices, dict):
            raise RuntimeError(f"模型 {model_name} 的官方价格必须是对象")

        parsed: dict[str, Decimal | None] = {}
        for field in PRICE_FIELDS:
            value = model_prices.get(field)
            if value is None:
                parsed[field] = None
                continue
            try:
                price = to_decimal(value, f"{model_name}.{field}")
            except ValueError as exc:
                raise RuntimeError(str(exc)) from exc
            if price <= 0:
                raise RuntimeError(f"{model_name}.{field} 必须大于 0 或为 null")
            parsed[field] = price

        if not any(value is not None for value in parsed.values()):
            raise RuntimeError(f"模型 {model_name} 至少需要填写一种官方价格")
        prices[model_name] = parsed
    return prices


def get_display_group_ratio(model: dict[str, Any], group_ratios: dict[str, Any]) -> Decimal:
    enabled_groups = model.get("enable_groups")
    if not isinstance(enabled_groups, list) or not enabled_groups:
        return Decimal("1")

    ratios: list[Decimal] = []
    for group in enabled_groups:
        if not isinstance(group, str) or group not in group_ratios:
            continue
        try:
            ratio = to_decimal(group_ratios[group], f"group_ratio.{group}")
        except ValueError:
            continue
        if ratio.is_finite():
            ratios.append(ratio)
    return min(ratios) if ratios else Decimal("1")


def parse_first_dynamic_tier(expression: str) -> dict[str, Decimal] | None:
    billing_expression = expression.split("|||", maxsplit=1)[0].strip()
    billing_expression = re.sub(r"^v\d+:", "", billing_expression, count=1)
    tier_match = TIER_PATTERN.search(billing_expression)
    if not tier_match:
        return None

    prices: dict[str, Decimal] = {}
    for variable, raw_value in COEFFICIENT_PATTERN.findall(tier_match.group(1)):
        field = DYNAMIC_FIELD_MAP[variable]
        value = Decimal(raw_value)
        if field not in prices and value > 0:
            prices[field] = value
    return prices


def calculate_prices(
    model: dict[str, Any],
    group_ratios: dict[str, Any],
    exchange_rate: Decimal,
) -> dict[str, Decimal | None]:
    if model.get("quota_type") != 0:
        return {field: None for field in PRICE_FIELDS}

    group_ratio = get_display_group_ratio(model, group_ratios)
    conversion = group_ratio * exchange_rate

    if model.get("billing_mode") == "tiered_expr" and model.get("billing_expr"):
        dynamic_prices = parse_first_dynamic_tier(str(model["billing_expr"]))
        if dynamic_prices is None:
            return {field: None for field in PRICE_FIELDS}
        return {
            field: dynamic_prices.get(field) * conversion
            if field in dynamic_prices
            else None
            for field in PRICE_FIELDS
        }

    model_ratio = to_decimal(model.get("model_ratio", 0), "model_ratio")
    completion_ratio = to_decimal(
        model.get("completion_ratio", 0), "completion_ratio"
    )
    input_price = model_ratio * Decimal("2") * conversion

    cache_ratio = model.get("cache_ratio")
    create_cache_ratio = model.get("create_cache_ratio")
    return {
        "input": input_price,
        "output": input_price * completion_ratio,
        "cache_read": input_price
        * to_decimal(cache_ratio, "cache_ratio")
        if cache_ratio is not None
        else None,
        "cache_write": input_price
        * to_decimal(create_cache_ratio, "create_cache_ratio")
        if create_cache_ratio is not None
        else None,
    }


def apply_multiplier(
    prices: dict[str, Decimal | None], multiplier: Decimal, has_official_price: bool
) -> dict[str, Decimal | None]:
    if not has_official_price:
        return prices
    return {
        field: value * multiplier if value is not None else None
        for field, value in prices.items()
    }


def format_discount(
    prices: dict[str, Decimal | None],
    official_prices: dict[str, Decimal | None] | None,
) -> str:
    if official_prices is None:
        return "-"

    actual_input = prices.get("input")
    official_input = official_prices.get("input")
    if actual_input is None or official_input is None or official_input <= 0:
        return "-"

    discount = (actual_input / official_input * Decimal("10")).quantize(
        Decimal("0.01"),
        rounding=ROUND_HALF_UP,
    )
    return f"{format(discount, 'f').rstrip('0').rstrip('.')}折"


def build_rows(
    models: list[dict[str, Any]],
    group_ratios: dict[str, Any],
    exchange_rate: Decimal,
    official_prices: dict[str, dict[str, Decimal | None]],
    multiplier: Decimal,
) -> list[
    tuple[
        str,
        Decimal | None,
        Decimal | None,
        Decimal | None,
        Decimal | None,
        str,
    ]
]:
    rows: list[
        tuple[
            str,
            Decimal | None,
            Decimal | None,
            Decimal | None,
            Decimal | None,
            str,
        ]
    ] = []
    for model in sorted(
        models,
        key=lambda item: str(item.get("model_name", "")).casefold(),
    ):
        model_name = str(model.get("model_name", "")).strip()
        if not model_name:
            continue
        model_official_prices = official_prices.get(model_name)
        prices = calculate_prices(model, group_ratios, exchange_rate)
        prices = apply_multiplier(
            prices,
            multiplier,
            model_official_prices is not None,
        )
        rows.append(
            (
                model_name.replace("\n", " "),
                prices["input"],
                prices["output"],
                prices["cache_read"],
                prices["cache_write"],
                format_discount(prices, model_official_prices),
            )
        )
    return rows


def write_excel(
    output: Path,
    rows: list[
        tuple[
            str,
            Decimal | None,
            Decimal | None,
            Decimal | None,
            Decimal | None,
            str,
        ]
    ],
) -> None:
    if output.suffix.lower() != ".xlsx":
        raise RuntimeError("输出文件必须使用 .xlsx 扩展名")
    if not output.parent.is_dir():
        raise RuntimeError(f"输出目录不存在：{output.parent}")

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    except ImportError as exc:
        raise RuntimeError("缺少 openpyxl，请先执行：pip install openpyxl") from exc

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "模型单价"
    worksheet.append(TABLE_HEADERS)

    for row in rows:
        worksheet.append(
            [
                row[0],
                *[
                    float(value) if value is not None else "-"
                    for value in row[1:5]
                ],
                row[5],
            ]
        )

    header_fill = PatternFill("solid", fgColor="1F4E78")
    data_fill = PatternFill("solid", fgColor="FFFFFF")
    thin_side = Side(style="thin", color="000000")
    cell_border = Border(
        left=thin_side,
        right=thin_side,
        top=thin_side,
        bottom=thin_side,
    )
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = cell_border

    for row in worksheet.iter_rows(min_row=2, min_col=1, max_col=6):
        for cell in row:
            cell.fill = data_fill
            cell.border = cell_border
            cell.alignment = Alignment(vertical="center")

    for row in worksheet.iter_rows(min_row=2, min_col=2, max_col=5):
        for cell in row:
            cell.number_format = "0.0#####"
            cell.alignment = Alignment(horizontal="right", vertical="center")

    worksheet.freeze_panes = "A2"
    worksheet.sheet_view.showGridLines = False
    worksheet.row_dimensions[1].height = 24
    width_limits = {
        "A": (28, 48),
        "B": (24, 30),
        "C": (24, 30),
        "D": (28, 34),
        "E": (28, 34),
        "F": (22, 32),
    }
    for column_index, column in enumerate(width_limits, start=1):
        values = [TABLE_HEADERS[column_index - 1]]
        values.extend(row[column_index - 1] for row in rows)
        display_width = max(
            sum(2 if ord(character) > 255 else 1 for character in str(value))
            for value in values
        )
        minimum_width, maximum_width = width_limits[column]
        worksheet.column_dimensions[column].width = max(
            minimum_width,
            min(display_width + 3, maximum_width),
        )

    first_note_row = len(rows) + 3
    for offset, note in enumerate(TABLE_NOTES):
        note_row = first_note_row + offset
        worksheet.merge_cells(
            start_row=note_row,
            start_column=1,
            end_row=note_row,
            end_column=6,
        )
        note_cell = worksheet.cell(row=note_row, column=1, value=note)
        note_cell.fill = data_fill
        note_cell.font = Font(color="000000")
        note_cell.alignment = Alignment(horizontal="left", vertical="center")

    try:
        workbook.save(output)
    except OSError as exc:
        raise RuntimeError(f"无法写入 Excel 文件 {output}：{exc}") from exc


def main() -> int:
    args = parse_args()
    try:
        official_prices = load_official_prices(args.official_prices)
        status_payload = request_json(
            args.base_url,
            "/api/status",
            args.token,
            float(args.timeout),
        )
        pricing_payload = request_json(
            args.base_url,
            "/api/pricing",
            args.token,
            float(args.timeout),
        )

        status = status_payload.get("data")
        models = pricing_payload.get("data")
        group_ratios = pricing_payload.get("group_ratio")
        if not isinstance(status, dict):
            raise RuntimeError("/api/status 返回的数据结构不正确")
        if not isinstance(models, list):
            raise RuntimeError("/api/pricing 返回的模型列表不正确")
        if not isinstance(group_ratios, dict):
            raise RuntimeError("/api/pricing 返回的分组倍率不正确")

        exchange_rate = to_decimal(
            status.get("usd_exchange_rate", 1),
            "usd_exchange_rate",
        )
        if exchange_rate <= 0:
            raise RuntimeError("usd_exchange_rate 必须大于 0")
        rows = build_rows(
            models,
            group_ratios,
            exchange_rate,
            official_prices,
            args.multiplier,
        )
        write_excel(args.output, rows)
    except (RuntimeError, ValueError) as exc:
        print(f"错误：{exc}", file=sys.stderr)
        return 1

    print(f"已输出 {len(rows)} 个模型到 {args.output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
